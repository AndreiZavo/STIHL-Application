import openpyxl

import xlrd

#xlsx_file = Path('D:\Barcode Software\STIHL Last App\Tabel_Clienti2.xlsx')
#
# wb = openpyxl.load_workbook('D:\Barcode Software\STIHL Last App\Tabel_Clienti2.xlsx')
# sheet = wb['Sheet1']
#
# book = Workbook()
# excel = book.active
#
# client = Client(name='Zavo')
#
# rows = (
#     (client.id, client.name, client.creation_date.__str__(), client.expiration_date.__str__()),
#     (89, 38, 12)
# )
#
# for row in rows:
#     excel.append(row)
#
# # book.save('Tabel_Clienti2.xlsx')
#
# a1 = excel['A1']
# a2 = excel.cell(row=1, column=3)
#
# print(a1.value)
# print(a2.value)
#
# col_names = []
# for column in excel.iter_cols(1, excel.max_column):
#     col_names.append(column[0].value)
#
#
# theFile = openpyxl.load_workbook('Tabel_Clienti2.xlsx')
# sheet = theFile[theFile.sheetnames[0]]
# print(sheet['A1'].value)
#
# sheet.insert_rows(2)
# #sheet.delete_rows(idx=2)
#
# theFile.save('Tabel_Clienti2.xlsx')

wb = openpyxl.load_workbook('Tabel_Clienti2.xlsx')
ws = wb.worksheets[0]

''' Add row '''
# client = Client(name='Zavo')
# client1 = Client(name='Ilinca')
# client2 = Client(name='Marcel')
#
# rows = (
#     (client.name, client.id, client.creation_date.__str__(), client.expiration_date.__str__()),
#     (client1.name, client1.id, client1.creation_date.__str__(), client1.expiration_date.__str__()),
#     (client2.name, client2.id, client2.creation_date.__str__(), client2.expiration_date.__str__())
# )
#
# for row in rows:
#     ws.append(row)
#
# wb.save('Tabel_Clienti2.xlsx')


''' Delete row '''
# for row in range(1, ws.max_row + 1):
#     for column in 'B':
#         cell_name = "{}{}".format(column, row)
#         if ws[cell_name].value == 'I2602202132':
#             ws.delete_rows(idx=row)
#
# wb.save('Tabel_Clienti2.xlsx')

import pandas as pd

df = pd.read_excel('Tabel_Clienti2.xlsx')
# print(df)

# print(df.columns)   # print headers

# print(df.iloc[0, 0])     # print specific cell

# for index, row in df.iterrows():
#    print(index, row)
#
workbook = xlrd.open_workbook('Tabel_Clienti2.xlsx')
# for sheet in workbook.sheets():
#     for row in range(sheet.nrows):
#         if sheet.cell(row, 1).value != '':
#             for column in range(sheet.ncols):
#                 print(sheet.cell(row, column).value, end=" ")
#             print('\n')

line = []
for sheet in workbook.sheets():
    for column in range(sheet.ncols):
        if sheet.cell(1, 1).value != '':
            line.append(sheet.cell(1, column).value)
