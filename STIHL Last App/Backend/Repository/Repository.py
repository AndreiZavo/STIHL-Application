import openpyxl
import xlrd

from Backend.Exceptions.Errors import RepositoryException


class Repository(object):

    def __init__(self):
        self._clients = []

    """
        Function returns a list containing all Client object   
        Input:  self
        Output: list
    """
    def get_all(self):
        return self._clients[:]

    """
        Function adds Client object to repository
        Input:  self, Client - client
        Output: 
    """
    def add_element(self, client):
        for index in range(len(self._clients)):
            if self._clients[index].id == client.id:
                raise RepositoryException("Duplicate client")
        self._clients.append(client)

    """
        Function deletes a Client object from repository   
        Input:  self, Client - client
        Output: bool
    """
    def remove_element(self, client):
        for index in range(len(self._clients)):
            if self._clients[index].name == client.name and self._clients[index].id == client.id:
                del self._clients[index]
                return
        raise RepositoryException("Client not found")


class FileRepository(Repository):

    def __init__(self, filename, read_element, write_element):
        self._filename = filename
        self._read_element = read_element
        self._write_element = write_element
        super().__init__()

    """
        Function reads text line, converts them to Client object and adds them to the list   
        Input:  self
        Output: 
    """
    def _read_from_file(self):
        self._clients = []
        workbook = xlrd.open_workbook(self._filename)
        for sheet in workbook.sheets():
            for row in range(sheet.nrows):
                line = []
                if sheet.cell(row, 1).value != '':
                    for column in range(sheet.ncols):
                        line.append(sheet.cell(row, column).value)
                    entity = self._read_element(line)
                    self._clients.append(entity)

    """
        Function writes all Client objects to the file
        Input:  self
        Output: 
    """
    def _write_new_entity_to_file(self):
        wb = openpyxl.load_workbook(self._filename)
        ws = wb.worksheets[0]
        line = self._write_element(self._clients[-1])
        ws.append(line)
        wb.save(self._filename)

    """
        Function overwrites the add_element function, reads and writes to the file    
        Input:  self
        Output: 
    """
    def add_element(self, client):
        self._read_from_file()
        Repository.add_element(self, client)
        self._write_new_entity_to_file()

    """
        Function returns a list of all Client object from a file
        Input: self
        Output: 
    """
    def get_all(self):
        self._read_from_file()
        return Repository.get_all(self)

    """
        Function deletes a specific client object from the list
        Input:  self
        Output: 
    """
    def remove_element(self, client):
        self._read_from_file()
        Repository.remove_element(self, client)
        self._remove_entity_from_file(client)

    '''
        Function deletes a specific row from file
        Input: self, client - Client
        Output: void
    '''
    def _remove_entity_from_file(self, client):
        wb = openpyxl.load_workbook(self._filename)
        ws = wb.worksheets[0]
        for row in range(1, ws.max_row + 1):
            for column in 'B':
                cell_name = "{}{}".format(column, row)
                if ws[cell_name].value == str(client.id):
                    ws.delete_rows(idx=row)
        wb.save(self._filename)
